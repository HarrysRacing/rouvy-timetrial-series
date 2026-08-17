#admin tool to load data from excel to the database in order to setup:
# Series
# Stages
# Races

# Admin tool will load the data based on worksheets
#The worksheets must be created as Series, Stages and Races

#Series and Stage dates/times will be entered in local timezone (EST) and should be converted to UTC for storage 

from pathlib import Path
from openpyxl import load_workbook

from datetime import datetime, timezone, UTC
from zoneinfo import ZoneInfo

from services.db_utils import DatabaseError, get_access_token, save_stage_route_data, save_race_data
from services.rouvy_api import get_rouvy_race_info

import sqlite3

# define the system ADMIN
ADMIN ="Old-Harry"

xlsx_path = Path(__file__).resolve().parent / "database" / "series_setup.xlsx"
db_path = Path(__file__).resolve().parent / "database" / "harrysracing.db"

# Load the Excel file
workbook = load_workbook(filename=xlsx_path)


try:
    
   #use "with" clause to manage transaction commit or rollback depending on success
   with sqlite3.connect(db_path) as conn:
      cursor = conn.cursor()


      # Series worksheets
      sheet = workbook["Series"]
      
      # Insert the data for the "Current" Series into the Series table
      for row in sheet.iter_rows(min_row=2):
         col_Status = row[0].value
         col_Name = row[1].value
         col_StartDate = row[2].value
         col_EndDate = row[3].value
         col_CountingStages = row[4].value
         col_LocalTZ = row[5].value
        
         if col_Status == 'Current':

           #set current Series name (for later use)
           currentSeries = col_Name
           localTZ = col_LocalTZ

           # Format dates to ISO 8601 with local TZ
           startDate_UTC = col_StartDate.replace(tzinfo=ZoneInfo(localTZ)).astimezone(UTC)
           endDate_UTC = col_EndDate.replace(tzinfo=ZoneInfo(localTZ)).astimezone(UTC)

           # Add ISO Format
           startDate_ISO = startDate_UTC.isoformat().replace("+00:00", "Z")       
           endDate_ISO = endDate_UTC.isoformat().replace("+00:00", "Z")       
        
           # check if Name exists, if so flag this
           query = ('SELECT * FROM Series WHERE Name = ?;')
       
           cursor.execute(query,(col_Name,))
           itExists = cursor.fetchall()
      
           if not itExists:
              query =  ( 'INSERT INTO Series '
                         '(Name, StartDate, EndDate, CountingStages) '
                         ' VALUES (?, ?, ?, ?); '
                        )
              cursor.execute(query, (col_Name, startDate_ISO, endDate_ISO, col_CountingStages))
           else:
              print('INSERT failed:', col_Name, ' already exists in the Series table.')
       
           break

      # Stages worksheet
      sheet = workbook["Stages"]

      # Insert the data for stages of the the "Current" Series into the Stage table
      for row in sheet.iter_rows(min_row=2):
         col_Series = row[0].value
         col_StgName = row[1].value
         col_StartDate = row[2].value
         col_EndDate = row[3].value
         col_RouteId = row[4].value
 
         if col_Series == currentSeries:

            # Format dates to ISO 8601 with local TZ
            startDate_UTC = col_StartDate.replace(tzinfo=ZoneInfo(localTZ)).astimezone(UTC)
            endDate_UTC = col_EndDate.replace(tzinfo=ZoneInfo(localTZ)).astimezone(UTC)

            # Add ISO Format
            startDate_ISO = startDate_UTC.isoformat().replace("+00:00", "Z")       
            endDate_ISO = endDate_UTC.isoformat().replace("+00:00", "Z")       
        
            # check if Stage name exists for the current Series, if so flag this
            query = ('SELECT * FROM Stage AS st, Series AS s '
                     'WHERE s.Id = st.SeriesId '
                     'AND s.Name = ? '
                     'AND st.Name = ?;'
                     )
       
            cursor.execute(query,(col_Series,col_StgName,))
            itExists = cursor.fetchall()
              
            if not itExists:
               query =  ( 'INSERT INTO Stage '
                          '(SeriesId, Name, StartDate, EndDate, RouteId) '
                          ' SELECT Id, ?, ?, ?, ? FROM Series WHERE Name = ?; '
                         )
               cursor.execute(query, (col_StgName, startDate_ISO, endDate_ISO, col_RouteId, col_Series))
            else:
               print('INSERT failed:', col_Name, ' already exists in the Stage table for the current Series.')


      # Have to access ROUVY API to get Event Info - first check OAuth Token
      
      
      #Get the RiderId for Old-Harry, me, the Admin...
      query = ('SELECT Id FROM Rider WHERE UserName = ?;')
      
      cursor.execute(query,(ADMIN,))
      riderId = cursor.fetchall()
      
      #Call function to get AccessToken
      accessToken = get_access_token(riderId[0])
      
      # Races worksheet
      sheet = workbook["Races"]

      # Insert the data for stages of the the "Current" Series into the Stage table
      for row in sheet.iter_rows(min_row=2):
         col_EventId = row[0].value
         
         if col_EventId:
            eventInfo = get_rouvy_race_info(accessToken, col_EventId) 
            
            #IF values do not already exist, update Stage : 
            # RouteId
            # RouteName
            # Country
            # Distance
            # Ascent
            # MaxSlope
            # WHERE StartDate of Race is within StartDate and EndDate of Stage
            # AND RouteId matches            
            
            routeId = eventInfo["event"]["route"]["routeId"]
            routeName = eventInfo["event"]["route"]["name"]
            country = eventInfo["event"]["route"]["countryCodeISO"]
            distance = (eventInfo["event"]["route"]["distanceMeters"]/1000) # to be saved in kms
            ascent = eventInfo["event"]["route"]["ascendedMeters"]
            maxSlope = eventInfo["event"]["route"]["slope"]["maxPercent"]
            raceStartUTC = eventInfo["event"]["startDateTimeUtc"]            
            raceName = eventInfo["event"]["title"]

            stageId = save_stage_route_data(routeId, routeName, country, distance, ascent, maxSlope, raceStartUTC)
            
            if stageId:
               save_race_data(stageId, col_EventId, raceStartUTC, raceName)                

 #CREATE TABLE Race (	
 # Id	INTEGER PRIMARY KEY,
 # StageId INTEGER, -- FK to Stage table
 # EventId	TEXT UNIQUE, -- FK to ROUVY Event infomation accessed via API
 # StartTime	TEXT, -- Date & Time for Race start -- ISO8601,'', 'YYYY-MM-DDTHH,'',MM,'',SSZ'
 # Name TEXT
 # );                       

except sqlite3.Error as error:
   raise DatabaseError("Unable to save Series/Stage data from Excel") from error